#!/usr/bin/env python3
"""Add a co-occurrence sheet to an existing coding matrix .xlsx.

Usage:
    python build_cooccurrence.py <matrix_xlsx>

Reads the "Binary Matrix" sheet, computes pairwise co-occurrence counts
(upper triangle), and adds a "Co-occurrence" sheet to the same workbook.

Co-occurrence cell (i, j) = number of interviews where both code i and
code j are present (both = 1).  Diagonal (i, i) = total count for code i.
Below-diagonal cells are left empty (upper triangle only).
"""

import argparse
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def build_cooccurrence(xlsx_path: str) -> dict:
    """Read binary matrix and add co-occurrence sheet.  Returns summary stats."""
    wb = openpyxl.load_workbook(xlsx_path)

    if "Binary Matrix" not in wb.sheetnames:
        print("Error: 'Binary Matrix' sheet not found in workbook.", file=sys.stderr)
        sys.exit(1)

    ws_binary = wb["Binary Matrix"]

    # Parse header to find code columns (skip Interview ID, Name, Side)
    headers = []
    for col in range(1, ws_binary.max_column + 1):
        val = ws_binary.cell(row=1, column=col).value
        if val:
            headers.append((col, str(val)))

    meta_cols = {"Interview ID", "Name", "Side"}
    code_cols = [(col, name) for col, name in headers if name not in meta_cols]
    code_names = [name for _, name in code_cols]
    code_col_indices = [col for col, _ in code_cols]

    # Find data rows (skip header, stop before summary rows)
    data_rows = []
    for row in range(2, ws_binary.max_row + 1):
        interview_id = ws_binary.cell(row=row, column=1).value
        side_val = ws_binary.cell(row=row, column=3).value
        # Summary rows have label in column 3 like "Total", "Consumer subtotal"
        if side_val and "subtotal" in str(side_val).lower():
            break
        if side_val and str(side_val).lower() == "total":
            break
        if interview_id is not None or ws_binary.cell(row=row, column=2).value is not None:
            data_rows.append(row)

    # Build binary vectors per code
    n_codes = len(code_names)
    vectors = []  # list of lists: [code_idx] -> list of 0/1 per interview
    for col_idx in code_col_indices:
        vec = []
        for row in data_rows:
            val = ws_binary.cell(row=row, column=col_idx).value
            vec.append(1 if val == 1 else 0)
        vectors.append(vec)

    # Compute co-occurrence matrix (upper triangle)
    cooccurrence = [[0] * n_codes for _ in range(n_codes)]
    for i in range(n_codes):
        for j in range(i, n_codes):
            count = sum(1 for k in range(len(data_rows)) if vectors[i][k] == 1 and vectors[j][k] == 1)
            cooccurrence[i][j] = count

    # Create or replace Co-occurrence sheet
    if "Co-occurrence" in wb.sheetnames:
        del wb["Co-occurrence"]
    ws_co = wb.create_sheet("Co-occurrence")

    # Styles
    header_font = Font(name="Arial", bold=True, size=9)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    diag_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    centre = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    body_font = Font(name="Arial", size=9)

    # Header row (code names across top)
    ws_co.cell(row=1, column=1, value="").font = header_font
    for j, name in enumerate(code_names):
        cell = ws_co.cell(row=1, column=j + 2, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)
        cell.border = thin_border

    # Row labels (code names down left side) + data
    for i, name in enumerate(code_names):
        label_cell = ws_co.cell(row=i + 2, column=1, value=name)
        label_cell.font = header_font
        label_cell.fill = header_fill
        label_cell.border = thin_border

        for j in range(n_codes):
            cell = ws_co.cell(row=i + 2, column=j + 2)
            cell.border = thin_border
            cell.alignment = centre
            cell.font = body_font

            if j < i:
                # Below diagonal: leave empty
                cell.value = ""
            elif j == i:
                # Diagonal: total for this code
                cell.value = cooccurrence[i][j]
                cell.fill = diag_fill
                cell.font = header_font
            else:
                # Upper triangle: co-occurrence count
                cell.value = cooccurrence[i][j]

    # Auto-fit first column
    max_label_len = max((len(name) for name in code_names), default=10)
    ws_co.column_dimensions["A"].width = min(max_label_len + 2, 30)
    # Code columns narrow
    for j in range(len(code_names)):
        col_letter = openpyxl.utils.get_column_letter(j + 2)
        ws_co.column_dimensions[col_letter].width = 4

    wb.save(xlsx_path)

    # Summary stats
    nonzero_pairs = sum(
        1 for i in range(n_codes) for j in range(i + 1, n_codes) if cooccurrence[i][j] > 0
    )
    total_pairs = n_codes * (n_codes - 1) // 2

    return {
        "n_codes": n_codes,
        "n_interviews": len(data_rows),
        "nonzero_pairs": nonzero_pairs,
        "total_possible_pairs": total_pairs,
        "density": round(nonzero_pairs / max(total_pairs, 1), 3),
    }


def main():
    parser = argparse.ArgumentParser(
        description="Add a co-occurrence sheet to an existing coding matrix .xlsx."
    )
    parser.add_argument("matrix_xlsx", help="Path to .xlsx file with 'Binary Matrix' sheet")
    args = parser.parse_args()

    import json
    summary = build_cooccurrence(args.matrix_xlsx)
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
