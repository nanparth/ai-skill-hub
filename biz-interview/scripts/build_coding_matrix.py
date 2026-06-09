#!/usr/bin/env python3
"""Build a binary coding matrix (.xlsx) from JSON coding data.

Usage:
    python build_coding_matrix.py <codings_json> <output_xlsx> [--codebook <codebook_md>]

Input JSON format (list of interview coding records):
[
  {
    "interview_id": "09",
    "interview_name": "Sophie",
    "side": "consumer",
    "codings": {
      "google-first": {"present": 1, "citations": [...], "flags": [...]},
      "research-paralysis": {"present": 0, "citations": [], "flags": []}
    }
  },
  ...
]

Output: .xlsx with a "Binary Matrix" sheet containing:
  - Header row: Interview ID, Name, Side, then one column per code
  - One row per interview with binary 0/1 values
  - Summary rows: total, demand-side subtotal, supply-side subtotal
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def extract_code_order_from_codebook(codebook_path: str) -> list[str]:
    """Extract code names in family order from a codebook .md file."""
    codes = []
    pattern = re.compile(r"^###\s+\d+\.\d+\s+(\S+)")
    with open(codebook_path, "r", encoding="utf-8") as f:
        for line in f:
            match = pattern.match(line.strip())
            if match:
                codes.append(match.group(1))
    return codes


def extract_all_codes_from_data(records: list[dict]) -> list[str]:
    """Extract all unique code names from coding records, sorted alphabetically."""
    codes = set()
    for record in records:
        codes.update(record.get("codings", {}).keys())
    return sorted(codes)


def build_matrix(records: list[dict], code_order: list[str], output_path: str) -> dict:
    """Build the binary coding matrix xlsx and return summary stats."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Binary Matrix"

    # Styles
    header_font = Font(name="Arial", bold=True, size=10)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    summary_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    centre = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    body_font = Font(name="Arial", size=10)

    # Header row
    meta_headers = ["Interview ID", "Name", "Side"]
    headers = meta_headers + code_order
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centre
        cell.border = thin_border

    # Data rows
    sides_seen = set()
    for row_idx, record in enumerate(records, 2):
        side = record.get("side", "unknown")
        sides_seen.add(side)
        codings = record.get("codings", {})

        ws.cell(row=row_idx, column=1, value=record.get("interview_id", "")).font = body_font
        ws.cell(row=row_idx, column=2, value=record.get("interview_name", "")).font = body_font
        ws.cell(row=row_idx, column=3, value=side).font = body_font

        for col_offset, code in enumerate(code_order):
            code_data = codings.get(code, {})
            val = code_data.get("present", 0) if isinstance(code_data, dict) else int(code_data)
            cell = ws.cell(row=row_idx, column=4 + col_offset, value=val)
            cell.alignment = centre
            cell.border = thin_border
            cell.font = body_font

    n_interviews = len(records)
    summary_start = n_interviews + 2

    # Summary row: totals
    summary_labels = ["Total"]
    sides_list = sorted(sides_seen)
    for side in sides_list:
        summary_labels.append(f"{side.title()} subtotal")

    for label_idx, label in enumerate(summary_labels):
        row = summary_start + label_idx
        ws.cell(row=row, column=1, value="").font = header_font
        ws.cell(row=row, column=2, value="").font = header_font
        ws.cell(row=row, column=3, value=label).font = header_font
        ws.cell(row=row, column=3).fill = summary_fill
        ws.cell(row=row, column=3).border = thin_border

        for col_offset, code in enumerate(code_order):
            col = 4 + col_offset
            if label == "Total":
                # Sum formula for the entire data range
                start_cell = ws.cell(row=2, column=col).coordinate
                end_cell = ws.cell(row=n_interviews + 1, column=col).coordinate
                formula = f"=SUM({start_cell}:{end_cell})"
            else:
                # SUMPRODUCT filtering by side
                side_name = sides_list[label_idx - 1]
                side_col_start = ws.cell(row=2, column=3).coordinate.replace(str(2), "")
                side_range = f"${side_col_start}$2:${side_col_start}${n_interviews + 1}"
                data_col = ws.cell(row=2, column=col).coordinate.replace(str(2), "")
                data_range = f"${data_col}$2:${data_col}${n_interviews + 1}"
                formula = f'=SUMPRODUCT(({side_range}="{side_name}")*({data_range}))'

            cell = ws.cell(row=row, column=col, value=formula)
            cell.alignment = centre
            cell.fill = summary_fill
            cell.border = thin_border
            cell.font = header_font

    # Auto-fit column widths (approximate)
    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, summary_start + len(summary_labels))
        )
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max(max_len + 2, 8), 30)

    wb.save(output_path)

    # Compute summary stats
    total_assignments = sum(
        code_data.get("present", 0) if isinstance(code_data, dict) else int(code_data)
        for record in records
        for code_data in record.get("codings", {}).values()
    )
    codes_per_interview = total_assignments / max(n_interviews, 1)

    code_totals = {}
    for code in code_order:
        code_totals[code] = sum(
            (rec.get("codings", {}).get(code, {}).get("present", 0)
             if isinstance(rec.get("codings", {}).get(code, {}), dict)
             else int(rec.get("codings", {}).get(code, 0)))
            for rec in records
        )
    top_codes = sorted(code_totals.items(), key=lambda x: x[1], reverse=True)[:10]

    return {
        "total_interviews": n_interviews,
        "total_codes": len(code_order),
        "total_assignments": total_assignments,
        "codes_per_interview_avg": round(codes_per_interview, 1),
        "top_10_codes": [{"code": c, "count": n} for c, n in top_codes],
        "sides": {s: sum(1 for r in records if r.get("side") == s) for s in sides_list},
    }


def main():
    parser = argparse.ArgumentParser(
        description="Build a binary coding matrix (.xlsx) from JSON coding data."
    )
    parser.add_argument("codings_json", help="Path to JSON file with coding records")
    parser.add_argument("output_xlsx", help="Path for output .xlsx file")
    parser.add_argument(
        "--codebook", help="Path to codebook .md file for code ordering (optional)"
    )
    args = parser.parse_args()

    with open(args.codings_json, "r", encoding="utf-8") as f:
        records = json.load(f)

    if not records:
        print("Error: no coding records found in input JSON.", file=sys.stderr)
        sys.exit(1)

    if args.codebook and Path(args.codebook).exists():
        code_order = extract_code_order_from_codebook(args.codebook)
        # Add any codes in data but not in codebook
        data_codes = extract_all_codes_from_data(records)
        for code in data_codes:
            if code not in code_order:
                code_order.append(code)
    else:
        code_order = extract_all_codes_from_data(records)

    summary = build_matrix(records, code_order, args.output_xlsx)
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
