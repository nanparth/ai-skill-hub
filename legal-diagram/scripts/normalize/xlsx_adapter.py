"""XLSX adapter: parse a spreadsheet into NormalizedDoc tables.

Heavy import (openpyxl) is LAZY: imported inside parse(), never at module top.
"""
from __future__ import annotations

from . import NormalizedDoc, DocBlock, DocTable, limit_value, mark_truncated


def _cell_str(value) -> str:
    """Coerce a cell value to a stripped string; None -> ''."""
    if value is None:
        return ""
    return str(value).strip()


def parse(src: str, sheets=None, **opts) -> NormalizedDoc:
    import openpyxl  # lazy heavy import

    doc = NormalizedDoc(source_format="xlsx")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    try:
        max_sheets = limit_value(opts, "max_xlsx_sheets")
        max_rows = limit_value(opts, "max_xlsx_rows_per_sheet")
        max_cells = limit_value(opts, "max_xlsx_cells_per_sheet")
        target_names = list(sheets if sheets else wb.sheetnames)
        if max_sheets and len(target_names) > max_sheets:
            target_names = target_names[:max_sheets]
            mark_truncated(doc, "XLSX_SHEET_LIMIT_REACHED")
        idx = 0
        for name in target_names:
            ws = wb[name]
            rows = []
            cell_count = 0
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if max_rows and row_idx >= max_rows:
                    mark_truncated(doc, "XLSX_ROW_LIMIT_REACHED")
                    break
                row_values = tuple(row)
                if max_cells and cell_count + len(row_values) > max_cells:
                    mark_truncated(doc, "XLSX_CELL_LIMIT_REACHED")
                    break
                rows.append(row_values)
                cell_count += len(row_values)

            # Heading block per sheet (sheet name).
            doc.blocks.append(
                DocBlock(
                    text=name,
                    block_type="heading",
                    idx=idx,
                    anchor=f"sheet:{name}",
                    parent_heading=None,
                )
            )
            idx += 1

            if not rows:
                doc.tables.append(
                    DocTable(headers=[], rows=[], anchor=f"sheet:{name}", caption=name)
                )
                continue

            headers = [_cell_str(c) for c in rows[0]]
            data_rows = [[_cell_str(c) for c in row] for row in rows[1:]]
            doc.tables.append(
                DocTable(
                    headers=headers,
                    rows=data_rows,
                    anchor=f"sheet:{name}",
                    caption=name,
                )
            )
    finally:
        wb.close()

    return doc
